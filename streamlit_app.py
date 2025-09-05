# streamlit_app.py
import os
import time
import random
from datetime import datetime, date, time as dtime, timedelta, timezone
from typing import Optional
from io import BytesIO
import zipfile

import streamlit as st
import pandas as pd
import altair as alt

from zoneinfo import ZoneInfo
TZ = ZoneInfo("Asia/Aqtobe")  # GMT+5 (замени при необходимости)


# ====== УТИЛЫ ВРЕМЕНИ ======
def local_day_bounds_to_utc(d: date) -> tuple[datetime, datetime]:
    start_local = datetime.combine(d, dtime.min).replace(tzinfo=TZ)
    end_local   = datetime.combine(d, dtime.max).replace(tzinfo=TZ)
    return start_local.astimezone(timezone.utc), end_local.astimezone(timezone.utc)

def _ensure_aware_utc(dt: datetime) -> datetime:
    if dt.tzinfo is None:
        return dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)

def week_bounds(d: date) -> tuple[date, date]:
    start = d - timedelta(days=d.weekday())
    end = start + timedelta(days=6)
    return start, end


# ====== БАЗОВЫЕ НАСТРОЙКИ ======
st.set_page_config(
    page_title="Система отслеживания и учёта картофеля",
    page_icon="🥔",
    layout="wide",
)

# ====== КЛЮЧИ (через .streamlit/secrets.toml или env) ======
SUPABASE_URL = st.secrets.get("SUPABASE_URL", os.getenv("SUPABASE_URL"))
SUPABASE_ANON_KEY = st.secrets.get("SUPABASE_ANON_KEY", os.getenv("SUPABASE_ANON_KEY"))
DEFAULT_BATCH = st.secrets.get("DEFAULT_BATCH", os.getenv("DEFAULT_BATCH", ""))  # удобно автоподставлять текущий запуск

USE_SUPABASE = bool(SUPABASE_URL and SUPABASE_ANON_KEY)
_sb = None
if USE_SUPABASE:
    try:
        from supabase import create_client
        _sb = create_client(SUPABASE_URL, SUPABASE_ANON_KEY)
    except Exception as e:
        st.warning(f"Не удалось инициализировать Supabase: {e}")
        USE_SUPABASE = False

# ====== ОФОРМЛЕНИЕ ======
st.markdown(
    """
    <style>
      .block-container { padding-top: 2.25rem; }
      .hdr { display:flex; justify-content:space-between; align-items:center; }
      .hdr h1 { margin:0; font-size:26px; font-weight:800; letter-spacing:.3px; }
      .hdr .sub { opacity:.8 }
      .hdr + .spacer { height: 10px; }
      hr { margin: 10px 0 22px 0; opacity:.25; }
      .stDateInput, .stNumberInput, .stTextInput { margin-bottom: .35rem; }
    </style>
    """,
    unsafe_allow_html=True,
)

def header(sub: str | None = None):
    st.markdown(
        f"<div class='hdr'><h1>Система отслеживания и учёта картофеля</h1>{f'<div class=\"sub\">{sub}</div>' if sub else ''}</div>",
        unsafe_allow_html=True
    )
    st.markdown('<div class="spacer"></div>', unsafe_allow_html=True)
    st.markdown("<hr/>", unsafe_allow_html=True)

def df_view(df: pd.DataFrame, caption: str = ""):
    if caption:
        st.caption(caption)
    st.dataframe(df, use_container_width=True)


# ====== СЕССИЯ/РОУТЕР ======
if "authed" not in st.session_state:
    st.session_state["authed"] = False
if "route" not in st.session_state:
    st.session_state["route"] = "app"  # сразу в приложение
if "day_picker" not in st.session_state:
    st.session_state["day_picker"] = date.today()

def go(page: str):
    st.session_state["route"] = page
    st.rerun()


# ====== ЧТЕНИЕ ДАННЫХ (кэш с TTL для почти realtime) ======
@st.cache_data(ttl=5)  # каждые ~5 сек будут подтягиваться свежие данные
def fetch_events(point: Optional[str], start_dt: datetime, end_dt: datetime, batch: Optional[str] = None) -> pd.DataFrame:
    """Возвращает ts (naive UTC), point, potato_id, width_cm, height_cm (+ batch)."""
    if not USE_SUPABASE:
        return pd.DataFrame(columns=["ts","point","potato_id","width_cm","height_cm","batch"])
    try:
        q = _sb.table("events").select("*").order("ts", desc=False)
        if point:
            q = q.eq("point", point)
        if batch:
            q = q.eq("batch", batch)
        start_iso = _ensure_aware_utc(start_dt).isoformat()
        end_iso   = _ensure_aware_utc(end_dt).isoformat()
        data = q.gte("ts", start_iso).lte("ts", end_iso).execute().data
        df = pd.DataFrame(data) if data else pd.DataFrame(columns=["ts","point","potato_id","width_cm","height_cm","batch"])
        if "ts" in df.columns:
            df["ts"] = pd.to_datetime(df["ts"], utc=True, errors="coerce")\
                          .dt.tz_convert("UTC").dt.tz_localize(None)
        for col in ("width_cm","height_cm"):
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")
        return df
    except Exception as e:
        st.warning(f"Ошибка чтения из Supabase: {e}")
        return pd.DataFrame(columns=["ts","point","potato_id","width_cm","height_cm","batch"])


# ====== АГРЕГАЦИИ ======
def hour_counts(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty or "ts" not in df.columns:
        return pd.DataFrame({"hour": [], "count": []})
    ts_local = (pd.to_datetime(df["ts"], utc=True, errors="coerce")
                  .dt.tz_localize("UTC").dt.tz_convert(TZ))
    hours_naive = ts_local.dt.floor("h").dt.tz_localize(None)
    return (
        pd.DataFrame({"hour": hours_naive, "potato_id": df["potato_id"]})
          .groupby("hour", as_index=False)
          .agg(count=("potato_id", "nunique"))
    )

def day_counts(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty or "ts" not in df.columns:
        return pd.DataFrame({"day": [], "count": []})
    ts_local = (pd.to_datetime(df["ts"], utc=True, errors="coerce")
                  .dt.tz_localize("UTC").dt.tz_convert(TZ))
    days_naive = ts_local.dt.floor("D").dt.tz_localize(None)
    return (
        pd.DataFrame({"day": days_naive, "potato_id": df["potato_id"]})
          .groupby("day", as_index=False)
          .agg(count=("potato_id", "nunique"))
    )


# ====== КАТЕГОРИИ ======
CATEGORIES = ["<30", "30–40", "40–50", "50–60", ">60"]

def bins_table(dfA: pd.DataFrame, dfB: pd.DataFrame) -> pd.DataFrame:
    """A=Изначально, B=Собрано. Возвращает таблицу по категориям."""
    def count_bins(df: pd.DataFrame) -> pd.Series:
        if df.empty or ("width_cm" not in df.columns):
            return pd.Series({c: 0 for c in CATEGORIES})
        bins = [0,30,40,50,60,10_000]
        labels = CATEGORIES
        cut = pd.cut(
            df["width_cm"].fillna(-1),
            bins=bins, labels=labels, right=False, include_lowest=True
        )
        vc = cut.value_counts().reindex(labels).fillna(0).astype(int)
        return vc

    A = count_bins(dfA)
    B = count_bins(dfB)
    losses = (A - B).clip(lower=0)
    loss_pct = pd.Series({c: (0.0 if A[c]==0 else round(losses[c]/A[c]*100, 1)) for c in CATEGORIES})

    return pd.DataFrame({
        "Категория":    CATEGORIES,
        "Изначально":   [int(A[c])      for c in CATEGORIES],
        "Потери (шт)":  [int(losses[c]) for c in CATEGORIES],
        "Собрано":      [int(B[c])      for c in CATEGORIES],
        "% потери":     [float(loss_pct[c]) for c in CATEGORIES],
    })


# ====== ДЕМО-ДАННЫЕ (оставим — вдруг пригодится оффлайн) ======
def demo_generate(day: date, base: int = 620, jitter: int = 90, seed: int = 42):
    rng = random.Random(seed + int(day.strftime("%Y%m%d")))
    hours = [datetime.combine(day, dtime(h,0)) for h in range(24)]
    rowsA, rowsB = [], []
    pid = 1
    for ts in hours:
        countA = max(0, int(rng.gauss(base, jitter)))
        countB = int(countA * rng.uniform(0.70, 0.85))
        for _ in range(countA):
            width = max(25.0, min(75.0, rng.gauss(52.0, 8.5)))
            rowsA.append({"ts": ts + timedelta(minutes=rng.randint(0,59)), "point":"A", "potato_id":pid, "width_cm":width, "height_cm":width*0.7})
            pid += 1
        for _ in range(countB):
            width = max(25.0, min(75.0, rng.gauss(53.0, 7.5)))
            rowsB.append({"ts": ts + timedelta(minutes=rng.randint(0,59)), "point":"B", "potato_id":pid, "width_cm":width, "height_cm":width*0.7})
            pid += 1
    return pd.DataFrame(rowsA), pd.DataFrame(rowsB)

def demo_generate_range(ref_day: date, days: int = 31):
    dfAs, dfBs = [], []
    for d in range(days-1, -1, -1):
        day_i = ref_day - timedelta(days=d)
        a, b = demo_generate(day_i)
        dfAs.append(a); dfBs.append(b)
    return pd.concat(dfAs, ignore_index=True), pd.concat(dfBs, ignore_index=True)


# ====== ЧАРТЫ ======
def render_hour_chart_grouped(dfA: pd.DataFrame, dfB: pd.DataFrame):
    ha = hour_counts(dfA).rename(columns={"count": "initial"})
    hb = hour_counts(dfB).rename(columns={"count": "collected"})
    merged = pd.merge(ha, hb, on="hour", how="outer").sort_values("hour")
    if merged.empty:
        st.info("Нет данных за выбранный период.")
        return pd.DataFrame()

    merged[["initial", "collected"]] = merged[["initial", "collected"]].fillna(0).astype(int)
    merged["diff"] = (merged["initial"] - merged["collected"]).clip(lower=0)

    long_df = pd.concat([
        merged[["hour", "collected"]].rename(columns={"collected": "Значение"}).assign(Сегмент="Итого (B)"),
        merged[["hour", "diff"]].rename(columns={"diff": "Значение"}).assign(Сегмент="Изначально (A)"),
    ], ignore_index=True)

    x_axis = alt.X(
        "hour:T",
        title="Дата и час",
        axis=alt.Axis(titlePadding=24, labelOverlap=True, labelFlush=True, titleAnchor="start"),
    )

    chart = (
        alt.Chart(long_df)
        .mark_bar()
        .encode(
            x=x_axis,
            y=alt.Y("Значение:Q", title="Количество", stack="zero"),
            color=alt.Color("Сегмент:N", title="", scale=alt.Scale(domain=["Итого (B)", "Изначально (A)"])),
            tooltip=[alt.Tooltip("hour:T", title="Час"), alt.Tooltip("Сегмент:N"), alt.Tooltip("Значение:Q", title="В сегменте")],
        )
        .properties(height=320, padding={"top": 10, "right": 12, "bottom": 44, "left": 8})
        .configure_axis(labelFontSize=12, titleFontSize=12)
    )

    st.altair_chart(chart, use_container_width=True)
    st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)
    return merged


# ====== ГЛАВНАЯ: Live + Daily ======
def page_dashboard_online():
    header()

    # ---- верхняя панель
    c1, c2, c3, c4 = st.columns([1.2, 1, 1, 1.2])

    with c1:
        # live-фильтр по batch (очень удобно отсеять старые тесты)
        batch_tag = st.text_input("batch-тег (рекомендуется)", value=DEFAULT_BATCH or "")
    with c2:
        auto_refresh = st.toggle("Автообновлять", value=True, help="Автоматически перезапускать с заданным интервалом")
    with c3:
        refresh_sec = st.number_input("Интервал, сек", min_value=2, max_value=60, value=5, step=1)
    with c4:
        if st.button("Обновить данные"):
            st.cache_data.clear()
            st.rerun()

    # статус подключения
    if USE_SUPABASE:
        try:
            _sb.table("events").select("potato_id").limit(1).execute()
            st.caption("✅ Подключение к Supabase: OK")
        except Exception as e:
            st.warning(f"⚠️ Supabase не отвечает: {e}")

    st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)

    # ==== LIVE РЕЖИМ (по умолчанию: последний час) ====
    st.subheader("Live: последние N минут")
    live_minutes = st.slider("Окно (мин)", min_value=5, max_value=180, value=60, step=5)
    now_utc = datetime.utcnow().replace(tzinfo=timezone.utc)
    live_start = now_utc - timedelta(minutes=live_minutes)

    dfA_live = fetch_events("A", live_start, now_utc, batch=batch_tag or None)
    dfB_live = fetch_events("B", live_start, now_utc, batch=batch_tag or None)

    live_init = dfA_live["potato_id"].nunique() if not dfA_live.empty else 0
    live_coll = dfB_live["potato_id"].nunique() if not dfB_live.empty else 0
    live_loss = max(0, live_init - live_coll)

    m1, m2, m3 = st.columns(3)
    m1.metric("Изначально (шт)", value=f"{live_init}")
    m2.metric("Потери (шт)", value=f"{live_loss}")
    m3.metric("Собрано (шт)", value=f"{live_coll}")

    # мини-таблички последних событий
    with st.expander("Последние события (A/B)"):
        colA, colB = st.columns(2)
        if not dfA_live.empty:
            dfA_tail = dfA_live.sort_values("ts", ascending=False).head(20)
            dfA_tail["ts"] = pd.to_datetime(dfA_tail["ts"]).dt.tz_localize("UTC").dt.tz_convert(TZ).dt.strftime("%Y-%m-%d %H:%M:%S")
            colA.write("A — последние 20:")
            df_view(dfA_tail[["ts","potato_id","width_cm","height_cm","batch"]])
        else:
            colA.info("Нет событий A за окно.")

        if not dfB_live.empty:
            dfB_tail = dfB_live.sort_values("ts", ascending=False).head(20)
            dfB_tail["ts"] = pd.to_datetime(dfB_tail["ts"]).dt.tz_localize("UTC").dt.tz_convert(TZ).dt.strftime("%Y-%m-%d %H:%M:%S")
            colB.write("B — последние 20:")
            df_view(dfB_tail[["ts","potato_id","width_cm","height_cm","batch"]])
        else:
            colB.info("Нет событий B за окно.")

    st.markdown("<hr/>", unsafe_allow_html=True)

    # ==== ЕЖЕДНЕВНЫЙ РАЗДЕЛ (по выбранной локальной дате) ====
    st.subheader("День (локальная дата)")
    dcol1, dcol2 = st.columns([1, 1])
    with dcol1:
        st.date_input("Дата", key="day_picker")
    with dcol2:
        st.caption(f"Часовой пояс: {TZ}")

    day = st.session_state["day_picker"]
    start_day_utc, end_day_utc = local_day_bounds_to_utc(day)

    dfA = fetch_events("A", start_day_utc, end_day_utc, batch=batch_tag or None)
    dfB = fetch_events("B", start_day_utc, end_day_utc, batch=batch_tag or None)

    # метрики за день
    total_initial = dfA["potato_id"].nunique() if not dfA.empty else 0
    total_collected = dfB["potato_id"].nunique() if not dfB.empty else 0
    total_losses = max(0, total_initial - total_collected)

    dm1, dm2, dm3 = st.columns(3)
    dm1.metric("Изначально (шт)", value=f"{total_initial}")
    dm2.metric("Потери (шт)", value=f"{total_losses}")
    dm3.metric("Собрано (шт)", value=f"{total_collected}")

    # поток по часам
    st.markdown("### Поток по часам")
    merged_hours = render_hour_chart_grouped(dfA, dfB)

    # Excel-отчёт (часы + категории)
    ha = hour_counts(dfA).rename(columns={"count": "Изначально (A)"})
    hb = hour_counts(dfB).rename(columns={"count": "Итого (B)"})
    hour_export = pd.merge(ha, hb, on="hour", how="outer").sort_values("hour").fillna(0)
    hour_export = hour_export.rename(columns={"hour": "Дата и час"})

    bins_df = bins_table(dfA, dfB)

    file_bytes, ext, mime = make_excel_bytes(hour_export, bins_df)
    st.download_button(
        label="Скачать отчёт" + (" (Excel)" if ext == "xlsx" else " (ZIP/CSV)"),
        data=file_bytes,
        file_name=f"potato_report_{day.isoformat()}." + ext,
        mime=mime,
        use_container_width=True
    )

    # Автоперезапуск live-панели
    if auto_refresh:
        time.sleep(refresh_sec)
        st.rerun()


# ====== УТИЛИТА: Excel-выгрузка ======
def make_excel_bytes(hour_df: pd.DataFrame, bins_df: pd.DataFrame) -> tuple[bytes, str, str]:
    try:
        import xlsxwriter  # noqa: F401
        buf = BytesIO()
        with pd.ExcelWriter(buf, engine="xlsxwriter", datetime_format="yyyy-mm-dd hh:mm") as writer:
            hour_df.to_excel(writer, index=False, sheet_name="Поток по часам")
            bins_df.to_excel(writer, index=False, sheet_name="Категории")
            wb = writer.book
            ws_hours = writer.sheets["Поток по часам"]
            ws_bins  = writer.sheets["Категории"]
            dt_fmt = wb.add_format({"num_format": "yyyy-mm-dd hh:mm"})
            for col_idx, col_name in enumerate(hour_df.columns):
                col_data = hour_df[col_name]
                if pd.api.types.is_datetime64_any_dtype(col_data):
                    width = max(len(str(col_name)), 16) + 2
                    ws_hours.set_column(col_idx, col_idx, width, dt_fmt)
                else:
                    max_len = max(len(str(col_name)), int(col_data.astype(str).map(len).max() or 0))
                    ws_hours.set_column(col_idx, col_idx, max_len + 2)
            for col_idx, col_name in enumerate(bins_df.columns):
                col_data = bins_df[col_name]
                max_len = max(len(str(col_name)), int(col_data.astype(str).map(len).max() or 0))
                ws_bins.set_column(col_idx, col_idx, max_len + 2)
        return buf.getvalue(), "xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    except Exception:
        pass

    try:
        import openpyxl  # noqa: F401
        from openpyxl.utils import get_column_letter
        buf = BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl", datetime_format="yyyy-mm-dd hh:mm") as writer:
            hour_df.to_excel(writer, index=False, sheet_name="Поток по часам")
            bins_df.to_excel(writer, index=False, sheet_name="Категории")
            ws_hours = writer.sheets["Поток по часам"]
            ws_bins  = writer.sheets["Категории"]
            def autofit(ws, df):
                for idx, col_name in enumerate(df.columns, start=1):
                    col_letter = get_column_letter(idx)
                    col_series = df[col_name]
                    if pd.api.types.is_datetime64_any_dtype(col_series):
                        width = max(len(str(col_name)), 16) + 2
                        ws.column_dimensions[col_letter].width = width
                        for row in range(2, len(col_series) + 2):
                            ws[f"{col_letter}{row}"].number_format = "yyyy-mm-dd hh:mm"
                    else:
                        max_len = max(len(str(col_name)), int(col_series.astype(str).map(len).max() or 0))
                        ws.column_dimensions[col_letter].width = max_len + 2
            # типы могут быть уже строками — не страшно
            autofit(ws_hours, hour_df); autofit(ws_bins, bins_df)
        return buf.getvalue(), "xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    except Exception:
        pass

    buf = BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("Поток по часам.csv", hour_df.to_csv(index=False))
        zf.writestr("Категории.csv",    bins_df.to_csv(index=False))
    return buf.getvalue(), "zip", "application/zip"


# ====== ЛОГИН (минимальный; если USE_SUPABASE=True) ======
def page_login():
    st.subheader("Вход в систему")
    with st.form("login_form", clear_on_submit=False):
        email = st.text_input("E-mail", placeholder="you@company.com")
        password = st.text_input("Пароль", type="password", placeholder="••••••••")
        submitted = st.form_submit_button("Войти")
    if submitted:
        ok = True
        if USE_SUPABASE:
            try:
                resp = _sb.auth.sign_in_with_password({"email": email, "password": password})
                ok = bool(getattr(resp, "user", None))
                if not ok:
                    st.error("Неверная почта или пароль.")
            except Exception as e:
                st.error(f"Ошибка авторизации: {e}")
                ok = False
        if ok:
            st.session_state["authed"] = True
            st.session_state["route"] = "app"
            st.rerun()
    st.caption("Доступ выдаётся администраторами.")


# ====== APP ======
def page_app():
    page_dashboard_online()


# ====== MAIN ======
def main():
    if not USE_SUPABASE:
        st.session_state["authed"] = True
        st.session_state["route"] = "app"

    route = st.session_state.get("route", "app")
    authed = st.session_state.get("authed", False)

    if route == "login" and USE_SUPABASE and not authed:
        header("Авторизация")
        page_login()
        return

    page_app()


if __name__ == "__main__":
    main()
